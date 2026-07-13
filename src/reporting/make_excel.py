import numpy as np, pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import os
OUT='../../outputs/'

wb=Workbook()

# ── helpers ──
def hdr(ws,row,col,val,bold=True,bg='1F77B4',fg='FFFFFF',sz=11):
    c=ws.cell(row,col,val)
    c.font=Font(bold=bold,color=fg,size=sz,name='Arial')
    c.fill=PatternFill('solid',start_color=bg)
    c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    return c
def cell(ws,row,col,val,fmt=None,bold=False,bg=None,align='center'):
    c=ws.cell(row,col,val)
    c.font=Font(bold=bold,name='Arial',size=10)
    c.alignment=Alignment(horizontal=align,vertical='center')
    if fmt: c.number_format=fmt
    if bg: c.fill=PatternFill('solid',start_color=bg)
    return c
def border_range(ws,r1,c1,r2,c2):
    thin=Side(style='thin')
    for r in range(r1,r2+1):
        for c in range(c1,c2+1):
            ws.cell(r,c).border=Border(left=thin,right=thin,top=thin,bottom=thin)

# ── Sheet 1: Spatial Metrics ──────────────────────────────────────────────────
ws1=wb.active; ws1.title='Spatial Metrics'
ws1.column_dimensions['A'].width=16
for col,w in zip('BCDEFG',[12,12,12,12,12,14]): ws1.column_dimensions[col].width=w
ws1.row_dimensions[1].height=30; ws1.row_dimensions[2].height=22

hdr(ws1,1,1,'Spatial Evaluation — Pseudo-Absence Methods (n=5500)',bg='1F77B4',sz=13)
ws1.merge_cells('A1:G1')
titles=['Method','K-func SSE↓','Centroid Dist↓','Grid Variance↓','Mean NN Dist','Fire NN Ref','Best?']
bgs=['2C2C2C','4472C4','4472C4','4472C4','4472C4','4472C4','4472C4']
for j,t in enumerate(titles): hdr(ws1,2,j+1,t,bg=bgs[j],sz=10)

sp=pd.read_csv(OUT+'spatial_results.csv')
method_colors={'Heuristic':'D6E4F7','Random':'D6F5D6','SA':'FFF0CC','GAN':'EFD6FF'}
rows_data=[(sp.loc[sp.Method==m].iloc[0]) for m in ['Heuristic','Random','SA','GAN']]
metrics=['K_SSE','Centroid','Grid_Var','Mean_NN']
best={'K_SSE':min,'Centroid':min,'Grid_Var':min,'Mean_NN':lambda v:min(v,key=lambda x:abs(x-0.0112))}

for r,row in enumerate(rows_data,3):
    bg=method_colors.get(row.Method,'FFFFFF')
    cell(ws1,r,1,row.Method,bold=True,bg=bg,align='left')
    for c,met in enumerate(metrics,2):
        v=float(row[met]); cell(ws1,r,c,v,fmt='0.0000',bg=bg)
    cell(ws1,r,6,float(row['Fire_NN_ref']),fmt='0.0000',bg='F2F2F2')
    # best marker
    vals_met={m2:float(sp.loc[sp.Method==m2,metrics[0]].values[0]) for m2 in ['Heuristic','Random','SA','GAN']}
    best_m=min(vals_met,key=vals_met.get)
    cell(ws1,r,7,'★ Best K-SSE' if row.Method==best_m else '',bold=True,bg=bg)
border_range(ws1,1,1,4+len(rows_data),7)

ws1['A9']='Note: K-func SSE = sum of squared deviations of K-function from fire reference. Lower = more fire-like spatial distribution.'
ws1['A9'].font=Font(italic=True,size=9,name='Arial')

# ── Sheet 2: ML Results ───────────────────────────────────────────────────────
ws2=wb.create_sheet('ML Results')
for col,w in zip('ABCDEFGHIJ',[14,10,10,10,10,10,10,10,10,12]): ws2.column_dimensions[col].width=w
ws2.row_dimensions[1].height=30; ws2.row_dimensions[2].height=40

hdr(ws2,1,1,'ML Evaluation — AUC & TSS (Spatial Block CV, 5-fold, n=5500)',bg='1F77B4',sz=13)
ws2.merge_cells('A1:J1')
headers2=['Model','AUC\nHeuristic','AUC\nRandom','AUC\nSA','AUC\nGAN',
          'TSS\nHeuristic','TSS\nRandom','TSS\nSA','TSS\nGAN','Best AUC Method']
for j,h in enumerate(headers2): hdr(ws2,2,j+1,h,bg='2C2C2C',sz=10)

cvH=np.load(OUT+'cv_h.npy',allow_pickle=True).item(); rH=cvH['res']
cvR=np.load(OUT+'cv_r.npy',allow_pickle=True).item(); rR=cvR['res']
cvS=np.load(OUT+'cv_s.npy',allow_pickle=True).item(); rS=cvS['res']
cvG=np.load(OUT+'cv_g.npy',allow_pickle=True).item(); rG=cvG['res']
ML=list(rH.keys())

model_bgs={'RandomForest':'D6E4F7','XGBoost':'FFF0CC','KNN':'D6F5D6','SVM':'FFE4E4'}
for r,nm in enumerate(ML,3):
    bg=model_bgs.get(nm,'FFFFFF')
    cell(ws2,r,1,nm,bold=True,bg=bg,align='left')
    aucs={'Heuristic':rH[nm]['AUC'],'Random':rR[nm]['AUC'],'SA':rS[nm]['AUC'],'GAN':rG[nm]['AUC']}
    tsss={'Heuristic':rH[nm]['TSS'],'Random':rR[nm]['TSS'],'SA':rS[nm]['TSS'],'GAN':rG[nm]['TSS']}
    best_m=max(aucs,key=aucs.get)
    for c,key in enumerate(['Heuristic','Random','SA','GAN'],2):
        v=aucs[key]; hl='FFC000' if key==best_m else bg
        cell(ws2,r,c,v,fmt='0.0000',bg=hl,bold=(key==best_m))
    for c,key in enumerate(['Heuristic','Random','SA','GAN'],6):
        cell(ws2,r,c,tsss[key],fmt='0.000',bg=bg)
    cell(ws2,r,10,f'★ {best_m}',bold=True,bg='FFC000')
border_range(ws2,1,1,2+len(ML),10)

# ── Sheet 3: Bootstrap ΔAUC ───────────────────────────────────────────────────
ws3=wb.create_sheet('Delta AUC Bootstrap')
for col,w in zip('ABCDE',[14,12,12,12,14]): ws3.column_dimensions[col].width=w
hdr(ws3,1,1,'Bootstrap ΔAUC 95% CI (2000 resamples) — Heuristic vs Random, GAN vs Random',bg='1F77B4',sz=12)
ws3.merge_cells('A1:E1')
for j,h in enumerate(['Model','ΔAUC H−R','95% CI H−R','ΔAUC GAN−R','95% CI GAN−R'],1):
    hdr(ws3,2,j,h,bg='2C2C2C',sz=10)

def bdelta(a,b,n=2000):
    np.random.seed(42)
    d=np.array([np.mean(np.random.choice(a,len(a),True))-np.mean(np.random.choice(b,len(b),True)) for _ in range(n)])
    return float(np.mean(d)),float(np.percentile(d,2.5)),float(np.percentile(d,97.5))

for r,nm in enumerate(ML,3):
    dm_hr,lo_hr,hi_hr=bdelta(rH[nm]['folds'],rR[nm]['folds'])
    dm_gr,lo_gr,hi_gr=bdelta(rG[nm]['folds'],rR[nm]['folds'])
    cell(ws3,r,1,nm,bold=True,bg='F2F2F2',align='left')
    cell(ws3,r,2,dm_hr,fmt='+0.0000;-0.0000',bg='D6F5D6' if dm_hr>0 else 'FFE4E4',bold=True)
    cell(ws3,r,3,f'[{lo_hr:+.4f}, {hi_hr:+.4f}]',bg='F9F9F9',align='center')
    cell(ws3,r,4,dm_gr,fmt='+0.0000;-0.0000',bg='D6F5D6' if dm_gr>0 else 'FFE4E4',bold=True)
    cell(ws3,r,5,f'[{lo_gr:+.4f}, {hi_gr:+.4f}]',bg='F9F9F9',align='center')
border_range(ws3,1,1,2+len(ML),5)

# ── Sheet 4: Figures ──────────────────────────────────────────────────────────
ws4=wb.create_sheet('Figures')
ws4['A1']='All Figures — PA Methods Comparison'
ws4['A1'].font=Font(bold=True,size=14,name='Arial')
figs=[('fig01_spatial_scatter.png','Fig 1: Spatial Distribution',2),
      ('fig02_ksse_bar.png','Fig 2: K-function SSE',35),
      ('fig03_spatial_metrics.png','Fig 3: Spatial Metrics Grid',65),
      ('fig04_fire_dist.png','Fig 4: Distance to Fire Distribution',95),
      ('fig05_auc_bar.png','Fig 5: AUC Comparison',125),
      ('fig06_tss_bar.png','Fig 6: TSS Comparison',155),
      ('fig07_roc.png','Fig 7: ROC Curves',185),
      ('fig08_delta_auc.png','Fig 8: ΔAUC Bootstrap CI',215),
      ('fig10_obs_pred.png','Fig 10: Observed vs Predicted',245),
      ('fig12_density_hex.png','Fig 12: Spatial Density',310),
      ('fig13_auc_heatmap.png','Fig 13: AUC Heatmap',340),
      ('fig14_tss_heatmap.png','Fig 14: TSS Heatmap',367)]
for fname,cap,anchor_row in figs:
    fpath=os.path.join(OUT,fname)
    if os.path.exists(fpath):
        ws4.cell(anchor_row,1,cap).font=Font(bold=True,size=10,name='Arial')
        img=XLImage(fpath); img.width=750; img.height=320
        ws4.add_image(img,f'A{anchor_row+1}')

wb.save(OUT+'PA_Methods_Results.xlsx')
print('Excel saved.')
